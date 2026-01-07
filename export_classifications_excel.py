#!/usr/bin/env python3
"""
Export Classification Results to Excel
Generates comprehensive Excel report with multiple worksheets
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_excel_report(json_file='qa_multidirectional_report.json', output_file='classification_report.xlsx'):
    """Create comprehensive Excel report from QA analysis"""

    print(f"📊 Loading data from {json_file}...")

    # Load JSON data
    with open(json_file, 'r') as f:
        data = json.load(f)

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. SUMMARY SHEET
    print("  Creating Summary sheet...")
    ws_summary = wb.create_sheet("Summary")

    ws_summary['A1'] = "CLASSIFICATION ANALYSIS SUMMARY"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:D1')

    ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary['A2'].font = Font(italic=True)
    ws_summary.merge_cells('A2:D2')

    # Overall metrics
    row = 4
    ws_summary[f'A{row}'] = "OVERALL METRICS"
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    ws_summary.merge_cells(f'A{row}:B{row}')

    row += 1
    metrics = [
        ("Total Queries Tested", data['summary']['total_queries']),
        ("Classified Queries", data['summary']['bottom_to_top']['classified']),
        ("Classification Rate", f"{data['summary']['bottom_to_top']['rate']:.1f}%"),
        ("Average Confidence", f"{data['summary']['confidence']['average']:.2f}"),
        ("Confidence Range", f"{data['summary']['confidence']['min']:.2f} - {data['summary']['confidence']['max']:.2f}"),
        ("Low Confidence Count", f"{data['summary']['confidence']['low_count']} queries"),
    ]

    for metric, value in metrics:
        ws_summary[f'A{row}'] = metric
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'] = value
        row += 1

    # Validation results
    row += 1
    ws_summary[f'A{row}'] = "VALIDATION RESULTS"
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    ws_summary.merge_cells(f'A{row}:C{row}')

    row += 1
    ws_summary[f'A{row}'] = "Test Type"
    ws_summary[f'B{row}'] = "Valid"
    ws_summary[f'C{row}'] = "Success Rate"

    for col in ['A', 'B', 'C']:
        ws_summary[f'{col}{row}'].fill = header_fill
        ws_summary[f'{col}{row}'].font = header_font
        ws_summary[f'{col}{row}'].border = thin_border

    row += 1
    validation_data = [
        ("Bottom-to-Top (Hierarchy)", data['summary']['bottom_to_top']['valid'],
         data['summary']['bottom_to_top']['classified']),
        ("Middle-to-Top (Rollup)", data['summary']['middle_to_top']['valid'],
         data['summary']['middle_to_top']['classified']),
        ("Middle-to-Bottom (Drilldown)", data['summary']['middle_to_bottom']['valid'],
         data['summary']['middle_to_bottom']['classified']),
    ]

    for test_type, valid, total in validation_data:
        ws_summary[f'A{row}'] = test_type
        ws_summary[f'B{row}'] = f"{valid}/{total}"
        rate = (valid / total * 100) if total > 0 else 0
        ws_summary[f'C{row}'] = f"{rate:.1f}%"

        # Color code based on success rate
        if rate == 100:
            ws_summary[f'C{row}'].fill = success_fill
        elif rate >= 80:
            ws_summary[f'C{row}'].fill = warning_fill
        else:
            ws_summary[f'C{row}'].fill = error_fill

        for col in ['A', 'B', 'C']:
            ws_summary[f'{col}{row}'].border = thin_border
        row += 1

    # Auto-size columns
    for col in ['A', 'B', 'C', 'D']:
        ws_summary.column_dimensions[col].width = 25

    # 2. BOTTOM-TO-TOP SHEET
    print("  Creating Bottom-to-Top sheet...")
    ws_bottom = wb.create_sheet("Bottom-to-Top")

    headers = ["Query ID", "Query", "L5 Keyword", "L4 Topic", "L3 Intent",
               "L2 Subcategory", "L1 Category", "Confidence", "Score", "Valid"]

    for col_num, header in enumerate(headers, 1):
        cell = ws_bottom.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for idx, item in enumerate(data['bottom_to_top'], 2):
        ws_bottom.cell(row=idx, column=1, value=item['query_id'])
        ws_bottom.cell(row=idx, column=2, value=item['query'])

        if item.get('classification'):
            c = item['classification']
            ws_bottom.cell(row=idx, column=3, value=c['L5'].get('keyword', c['L5'].get('topic', 'N/A')))
            ws_bottom.cell(row=idx, column=4, value=c['L4'].get('topic', 'N/A'))
            ws_bottom.cell(row=idx, column=5, value=c['L3'].get('intent_category', 'N/A'))
            ws_bottom.cell(row=idx, column=6, value=c['L2'].get('name', 'N/A'))
            ws_bottom.cell(row=idx, column=7, value=c['L1'].get('name', 'N/A'))
            ws_bottom.cell(row=idx, column=8, value=round(item.get('confidence', 0), 2))
            ws_bottom.cell(row=idx, column=9, value=item.get('score', 0))

            valid_cell = ws_bottom.cell(row=idx, column=10, value="✅" if item.get('hierarchy_valid') else "❌")
            valid_cell.fill = success_fill if item.get('hierarchy_valid') else error_fill
            valid_cell.alignment = center_align

            # Color code confidence
            conf = item.get('confidence', 0)
            conf_cell = ws_bottom.cell(row=idx, column=8)
            if conf >= 15:
                conf_cell.fill = success_fill
            elif conf >= 5:
                conf_cell.fill = warning_fill
            else:
                conf_cell.fill = error_fill
        else:
            for col in range(3, 11):
                ws_bottom.cell(row=idx, column=col, value="UNCLASSIFIED")
                ws_bottom.cell(row=idx, column=col).fill = error_fill

        # Apply borders
        for col in range(1, 11):
            ws_bottom.cell(row=idx, column=col).border = thin_border

    # Auto-size columns
    ws_bottom.column_dimensions['A'].width = 10
    ws_bottom.column_dimensions['B'].width = 35
    ws_bottom.column_dimensions['C'].width = 25
    ws_bottom.column_dimensions['D'].width = 30
    ws_bottom.column_dimensions['E'].width = 20
    ws_bottom.column_dimensions['F'].width = 30
    ws_bottom.column_dimensions['G'].width = 20
    ws_bottom.column_dimensions['H'].width = 12
    ws_bottom.column_dimensions['I'].width = 10
    ws_bottom.column_dimensions['J'].width = 8

    # Enable text wrapping for query column
    for row in range(2, len(data['bottom_to_top']) + 2):
        ws_bottom.cell(row=row, column=2).alignment = wrap_align

    # 3. MIDDLE-TO-TOP SHEET
    print("  Creating Middle-to-Top sheet...")
    ws_mid_top = wb.create_sheet("Middle-to-Top")

    headers = ["Query ID", "Query", "L3 Intent", "L2 Subcategory", "L1 Category", "Valid"]

    for col_num, header in enumerate(headers, 1):
        cell = ws_mid_top.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for idx, item in enumerate(data['middle_to_top'], 2):
        ws_mid_top.cell(row=idx, column=1, value=item['query_id'])
        ws_mid_top.cell(row=idx, column=2, value=item['query'])
        ws_mid_top.cell(row=idx, column=3, value=item.get('l3_intent', 'UNCLASSIFIED'))
        ws_mid_top.cell(row=idx, column=4, value=item.get('l2_subcategory', 'UNCLASSIFIED'))
        ws_mid_top.cell(row=idx, column=5, value=item.get('l1_category', 'UNCLASSIFIED'))

        valid_cell = ws_mid_top.cell(row=idx, column=6, value="✅" if item.get('rollup_valid') else "❌")
        valid_cell.fill = success_fill if item.get('rollup_valid') else error_fill
        valid_cell.alignment = center_align

        for col in range(1, 7):
            ws_mid_top.cell(row=idx, column=col).border = thin_border

    ws_mid_top.column_dimensions['A'].width = 10
    ws_mid_top.column_dimensions['B'].width = 40
    ws_mid_top.column_dimensions['C'].width = 20
    ws_mid_top.column_dimensions['D'].width = 35
    ws_mid_top.column_dimensions['E'].width = 20
    ws_mid_top.column_dimensions['F'].width = 8

    for row in range(2, len(data['middle_to_top']) + 2):
        ws_mid_top.cell(row=row, column=2).alignment = wrap_align

    # 4. MIDDLE-TO-BOTTOM SHEET
    print("  Creating Middle-to-Bottom sheet...")
    ws_mid_bottom = wb.create_sheet("Middle-to-Bottom")

    headers = ["Query ID", "Query", "L3 Intent", "L4 Topic", "L5 Keyword", "Confidence", "Valid"]

    for col_num, header in enumerate(headers, 1):
        cell = ws_mid_bottom.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for idx, item in enumerate(data['middle_to_bottom'], 2):
        ws_mid_bottom.cell(row=idx, column=1, value=item['query_id'])
        ws_mid_bottom.cell(row=idx, column=2, value=item['query'])
        ws_mid_bottom.cell(row=idx, column=3, value=item.get('l3_intent', 'UNCLASSIFIED'))
        ws_mid_bottom.cell(row=idx, column=4, value=item.get('l4_topic', 'UNCLASSIFIED'))
        ws_mid_bottom.cell(row=idx, column=5, value=item.get('l5_keywords', 'UNCLASSIFIED'))
        ws_mid_bottom.cell(row=idx, column=6, value=round(item.get('confidence', 0), 2))

        valid_cell = ws_mid_bottom.cell(row=idx, column=7, value="✅" if item.get('drilldown_valid') else "❌")
        valid_cell.fill = success_fill if item.get('drilldown_valid') else error_fill
        valid_cell.alignment = center_align

        # Color code confidence
        conf = item.get('confidence', 0)
        conf_cell = ws_mid_bottom.cell(row=idx, column=6)
        if conf >= 15:
            conf_cell.fill = success_fill
        elif conf >= 5:
            conf_cell.fill = warning_fill
        else:
            conf_cell.fill = error_fill

        for col in range(1, 8):
            ws_mid_bottom.cell(row=idx, column=col).border = thin_border

    ws_mid_bottom.column_dimensions['A'].width = 10
    ws_mid_bottom.column_dimensions['B'].width = 40
    ws_mid_bottom.column_dimensions['C'].width = 20
    ws_mid_bottom.column_dimensions['D'].width = 35
    ws_mid_bottom.column_dimensions['E'].width = 30
    ws_mid_bottom.column_dimensions['F'].width = 12
    ws_mid_bottom.column_dimensions['G'].width = 8

    for row in range(2, len(data['middle_to_bottom']) + 2):
        ws_mid_bottom.cell(row=row, column=2).alignment = wrap_align

    # 5. RECOMMENDATIONS SHEET
    print("  Creating Recommendations sheet...")
    ws_rec = wb.create_sheet("Recommendations")

    ws_rec['A1'] = "ACTIONABLE RECOMMENDATIONS"
    ws_rec['A1'].font = Font(bold=True, size=14)
    ws_rec.merge_cells('A1:D1')

    headers = ["Priority", "Category", "Issue", "Action", "Impact"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_rec.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    row = 4
    for rec in data.get('recommendations', []):
        priority_cell = ws_rec.cell(row=row, column=1, value=rec['priority'])
        if rec['priority'] == 'CRITICAL':
            priority_cell.fill = error_fill
            priority_cell.font = Font(bold=True)
        elif rec['priority'] == 'HIGH':
            priority_cell.fill = warning_fill
            priority_cell.font = Font(bold=True)
        else:
            priority_cell.fill = success_fill

        ws_rec.cell(row=row, column=2, value=rec['category'])
        ws_rec.cell(row=row, column=3, value=rec['issue'])
        ws_rec.cell(row=row, column=4, value=rec['action'])
        ws_rec.cell(row=row, column=5, value=rec['impact'])

        for col in range(1, 6):
            ws_rec.cell(row=row, column=col).border = thin_border
            ws_rec.cell(row=row, column=col).alignment = wrap_align

        row += 1

    ws_rec.column_dimensions['A'].width = 12
    ws_rec.column_dimensions['B'].width = 15
    ws_rec.column_dimensions['C'].width = 35
    ws_rec.column_dimensions['D'].width = 50
    ws_rec.column_dimensions['E'].width = 30

    # 6. DETAILED METRICS SHEET
    print("  Creating Detailed Metrics sheet...")
    ws_metrics = wb.create_sheet("Detailed Metrics")

    if 'detailed_metrics' in data:
        dm = data['detailed_metrics']

        row = 1
        ws_metrics[f'A{row}'] = "DETAILED PERFORMANCE METRICS"
        ws_metrics[f'A{row}'].font = Font(bold=True, size=14)
        ws_metrics.merge_cells(f'A{row}:B{row}')

        row += 2

        # Confidence Distribution
        ws_metrics[f'A{row}'] = "CONFIDENCE DISTRIBUTION"
        ws_metrics[f'A{row}'].font = Font(bold=True, size=12)
        ws_metrics.merge_cells(f'A{row}:B{row}')
        row += 1

        if 'confidence' in dm and 'distribution' in dm['confidence']:
            for range_label, count in dm['confidence']['distribution'].items():
                ws_metrics[f'A{row}'] = f"Confidence {range_label}"
                ws_metrics[f'B{row}'] = count
                row += 1

        row += 1

        # Intent Distribution
        ws_metrics[f'A{row}'] = "INTENT DISTRIBUTION"
        ws_metrics[f'A{row}'].font = Font(bold=True, size=12)
        ws_metrics.merge_cells(f'A{row}:B{row}')
        row += 1

        if 'intent_distribution' in dm:
            for intent, count in dm['intent_distribution'].items():
                ws_metrics[f'A{row}'] = intent
                ws_metrics[f'B{row}'] = count
                row += 1

        row += 1

        # Category Distribution
        ws_metrics[f'A{row}'] = "CATEGORY DISTRIBUTION"
        ws_metrics[f'A{row}'].font = Font(bold=True, size=12)
        ws_metrics.merge_cells(f'A{row}:B{row}')
        row += 1

        if 'category_distribution' in dm:
            for category, count in dm['category_distribution'].items():
                ws_metrics[f'A{row}'] = category
                ws_metrics[f'B{row}'] = count
                row += 1

    ws_metrics.column_dimensions['A'].width = 30
    ws_metrics.column_dimensions['B'].width = 15

    # Save workbook
    print(f"\n💾 Saving Excel file: {output_file}...")
    wb.save(output_file)
    print(f"✅ Excel report created successfully!")

    # Print summary
    print(f"\n📊 Report Summary:")
    print(f"   Worksheets: 6")
    print(f"   - Summary: Overview and key metrics")
    print(f"   - Bottom-to-Top: {len(data['bottom_to_top'])} queries")
    print(f"   - Middle-to-Top: {len(data['middle_to_top'])} queries")
    print(f"   - Middle-to-Bottom: {len(data['middle_to_bottom'])} queries")
    print(f"   - Recommendations: {len(data.get('recommendations', []))} items")
    print(f"   - Detailed Metrics: Distribution analysis")

    return output_file


if __name__ == '__main__':
    import sys

    json_file = 'qa_multidirectional_report.json'
    output_file = 'classification_report.xlsx'

    if len(sys.argv) > 1:
        json_file = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]

    print("="*80)
    print("📊 CLASSIFICATION RESULTS - EXCEL EXPORT")
    print("="*80)
    print()

    try:
        result = create_excel_report(json_file, output_file)
        print(f"\n{'='*80}")
        print(f"✅ Export complete: {result}")
        print(f"{'='*80}")
    except FileNotFoundError as e:
        print(f"❌ Error: {e}")
        print(f"\nMake sure to run 'python3 qa_multidirectional.py' first to generate the report.")
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        import traceback
        traceback.print_exc()
